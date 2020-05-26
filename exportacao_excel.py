import pandas as pd
import re
from openpyxl import load_workbook
import difflib
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import dropbox
import os


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param cell_range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


# Importa as variáveis do arquivo de referências
path_referencias = r'/home/joao/PycharmProjects/Controle Financeiro/Documentos/Parametros.txt'
chave_acesso_dropbox = None
txt_ref = None

with open(path_referencias, 'r') as f:
    params = f.read().split('\n')

for param in params:
    if param.__contains__('chave_acesso_dropbox'):
        chave_acesso_dropbox = param.split(': ')[-1]
    elif param.__contains__('txt_ref'):
        txt_ref = param.split(': ')[-1]

assert chave_acesso_dropbox is not None
assert txt_ref is not None

# Importando o arquivo texto de referência (do Dropbox)
dbx = dropbox.Dropbox(chave_acesso_dropbox)
md, res = dbx.files_download(txt_ref)
gastos = res.content.decode("utf-8")

# Separando o texto em linhas e dividindo as colunas para transformar em dataframe
gastos = gastos.split('\n')
gastos_split = [re.split('\s*-\s*', gasto) for gasto in gastos]

# Criando o dataframe
headers = ['Data', 'Local', 'Valor', 'Categoria', 'Descrição']
df = pd.DataFrame(gastos_split, columns=headers)

# Substitui virgulas por pontos no campo 'Valor':
df['Valor'] = [data.replace(',', '.') for data in df['Valor']]

# Transformando em números caso haja alguma expressão do tipo x+y na coluna 'Valor'
df['Valor'] = df.eval(df['Valor'])

# Transformando a primeira letra em maiúscula no campo 'Conta'
df['Categoria'] = df['Categoria'].str.title()

# Removendo grupos com o padrão [v] (já foram incluídos na planilha previamente
patternDel = "(\[v\])"
filtro = df['Data'].str.contains(patternDel)
df = df[~filtro]

# Resetando os índices
df = df.reset_index(drop=True)

# Verificando campos que não estão identificados corretamente e substituindo pelos rótulos corretos

# Lista de categorias:
categorias = ['Contas Residenciais', 'Transporte', 'Compras', 'Desenvolvimento Pessoal', 'Lazer',
              'Independência Financeira',
              'Despesas de longo prazo', 'Poupança Transporte', 'Poupança Desenvolvimento', 'Poupança Lazer',
              'Poupança Despesas', 'Poupança Independência']

for categoria in df['Categoria']:
    if categoria not in categorias:
        ind = df[df['Categoria'] == categoria].index[0]
        repl = difflib.get_close_matches(categoria, categorias)[0]
        df.iloc[ind, 3] = repl

        print('Categoria \'' + str(categoria) + '\' foi substituída por \'' + str(repl) + '\'')

# Salvando as informações na planilha

root = Tk()
ftypes = [('Excel file', "*.xlsx")]
ttl = "Selecione planilha para inserção de dados"
dir1 = r'~/Documents/Planejamento financeiro/Planejamentos mensais'
root.fileName = askopenfilename(filetypes=ftypes, initialdir=dir1, title=ttl)
root.destroy()

# Define o mês e ano para jogar os gastos
arquivo = root.fileName.split('/')[-1]
mes = arquivo[14:16]
ano = arquivo[16:18]

# Transformando o campo 'Data' em datas
datas = df['Data'].tolist()
datas = [data.replace('[ ', '').replace(' ]', '') for data in datas]

mes_ano = '/' + mes + '/' + ano
datas_final = [data + mes_ano for data in datas]

df['Data'] = datas_final
df['Data'] = pd.to_datetime(df['Data'], dayfirst=True).dt.strftime('%d/%m/%Y')

# -------------------------------------- INSERINDO DADOS NO EXCEL -------------------------------------

book = load_workbook(root.fileName)
writer = pd.ExcelWriter(root.fileName, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
last_row = writer.sheets['Controle de gastos (Jão)'].max_row

df.to_excel(writer, "Controle de gastos (Jão)", startcol=0, startrow=last_row, index=False, header=False)

writer.sheets["Controle de gastos (Jão)"].column_dimensions['A'].width = 10.5

# Refazendo as bordas
ws1 = writer.sheets['Planejamentos (Jão)']
ws2 = writer.sheets['Controle de gastos (Jão)']

# Objetos de cada tipo de borda
border1 = Border(top=Side(border_style='medium', color='000000'),
                 bottom=Side(border_style='medium', color='000000'),
                 left=Side(border_style='medium', color='000000'),
                 right=Side(border_style='medium', color='000000'))

border2 = Border(top=Side(border_style='medium', color='000000'),
                 bottom=Side(border_style='medium', color='000000'),
                 left=Side(border_style='medium', color='000000'),
                 right=Side(border_style='thin', color='000000'))

border3 = Border(top=Side(border_style='medium', color='000000'),
                 bottom=Side(border_style='thin', color='000000'),
                 left=Side(border_style='thin', color='000000'),
                 right=Side(border_style='thin', color='000000'))

border4 = Border(top=Side(border_style='medium', color='000000'),
                 bottom=Side(border_style='thin', color='000000'),
                 left=Side(border_style='thin', color='000000'),
                 right=Side(border_style='medium', color='000000'))

# Reaplicando as bordas nas células
style_range(ws1, 'A1:H3', border1)

style_range(ws1, 'A5:B5', border1)

style_range(ws1, 'A12:A13', border2)
style_range(ws1, 'B12:C12', border3)
style_range(ws1, 'D12:E12', border3)
style_range(ws1, 'G12:H12', border4)

style_range(ws1, 'A19:B19', border1)
style_range(ws1, 'A24:B24', border1)

style_range(ws2, 'G3:G11', border1)
style_range(ws2, 'G13:G17', border1)

# Finaliza processo e salva planilha
writer.save()

# Abre excel com novos dados inseridos para conferência do usuário
os.system('start excel.exe "' + root.fileName.replace('/', '\\') + '"')